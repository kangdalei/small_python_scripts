# -*- coding: utf-8 -*-
# 配合Cad lisp脚本plot2pdf.LSP 将打印生成PDF文件列表写入excel表格中
# PDF文件名中携带了 图号#图名<_N数量_>_cad文件名前3位字符_打印日期等信息
# 其中<_N数量_> 可选项
# 本脚本将 图号 图名 数量三项信息提取出来, 分别写入excel表格不同列
# 方便后续处理.

import os
import re
from openpyxl import Workbook  # 处理excel xlsx 库


def split_file_name(file_name):
    # 把文件名分割成 图号 图名 数量，返回一个列表
    # 如果数量不存在，以字符串NONE代替
    split_file_name_list = []
    m = re.match(r'^([^#]+)#([^_]+).+.pdf$', file_name)
    split_file_name_list.append(m.group(1))  # 图号
    split_file_name_list.append(m.group(2))  # 图名

    m_n = re.match(r'^.+_N(\d+)_.+.pdf$', file_name)
    if m_n:  # 是否有数量
        split_file_name_list.append(int(m_n.group(1)))  # 转换为整数
    else:
        split_file_name_list.append('NONE')
    return split_file_name_list


dir_lists = os.listdir()  # 工作目录下文件名列表
pdf_filenames = []  # 用来存储pdf文件信息列表

for file_name in dir_lists:
    if re.match(r'.+#.+.pdf$', file_name):  # 有图号(#前)的pdf文件
        pdf_filenames.append(split_file_name(file_name))

for pdf_filename in pdf_filenames:
    print(pdf_filename)

wb = Workbook()  # 创建工作薄
ws = wb.active  # 活动工作表

pathname = os.path.split(os.getcwd())[-1]  # 当前工作目录名

excel_name = pathname + '料单.xlsx'  # 要保存的excel文件名

ws['A1'] = excel_name
ws['A2'] = '序号'
ws['B2'] = '图号'
ws['C2'] = '图名'
ws['D2'] = '数量'
ws['E2'] = '材料规格'
ws['F2'] = '备注'
ws['G2'] = '单重'
ws['H2'] = '总重'

row = 3
col = 2

for row in range(3, 3 + len(pdf_filenames)):
    for col in range(2, 5):
        ws.cell(row=row, column=col).value = pdf_filenames[row - 3][col - 2]
        # if col == 4:  # 第4列，数量列
        #     # 设置格式：格式 数字如果是字符串格式，无用
        #     # 数字如果是整数格式，也不用设置，所以注释掉
        #     ws.cell(row=row, column=col).number_format = 'General'

# 处理序号, 从1开始
serial_num = 1
row = 3
for row in range(3, 3 + len(pdf_filenames)):
    ws.cell(row=row, column=1).value = str(serial_num)
    serial_num += 1

wb.save(excel_name)  # 保存excel文件
print(excel_name + "已保存到当前目录")
print("done!")
